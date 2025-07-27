from openpyxl import load_workbook
import json


def excel_to_fixture_openpyxl(input_excel, output_json):
    wb = load_workbook(filename=input_excel)
    fixture_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get headers from first row
        headers = [cell.value for cell in ws[1]]

        # Process each data row
        for row in ws.iter_rows(min_row=2):
            fields = {}
            pk = None

            for header, cell in zip(headers, row):
                if header == "id":
                    pk = cell.value
                elif cell.value is not None:
                    # Handle different value types
                    if isinstance(cell.value, str) and cell.value.startswith("["):
                        try:
                            fields[header] = json.loads(cell.value)
                        except:
                            fields[header] = cell.value
                    else:
                        fields[header] = cell.value

            if pk is not None:
                fixture_data.append(
                    {
                        "model": sheet_name.replace(
                            "_", ".", 1
                        ),  # Convert back to app.Model
                        "pk": pk,
                        "fields": fields,
                    }
                )

    with open(output_json, "w") as f:
        json.dump(fixture_data, f, indent=2)

    print(f"Fixture JSON saved to {output_json}")


# Usage
excel_to_fixture_openpyxl("output.xlsx", "new_fixture.json")
